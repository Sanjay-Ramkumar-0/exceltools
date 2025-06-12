import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from tabulate import tabulate
import statistics


class ExcelFile:
    def __init__(self):
        self.wb = Workbook()
        self.sheet = self.wb.active
        self.file_path = None

    def create_excel(self, name):
        if not name.endswith(".xlsx"):
            name += ".xlsx"
        self.file_path = os.path.abspath(name)
        self.wb.save(self.file_path)
        return self

    def save(self):
        if self.file_path:
            self.wb.save(self.file_path)

    def save_as(self, new_path):
        if not new_path.endswith(".xlsx"):
            new_path += ".xlsx"
        self.file_path = os.path.abspath(new_path)
        self.wb.save(self.file_path)

    def rename_file(self, new_name):
        if not new_name.endswith(".xlsx"):
            new_name += ".xlsx"
        new_path = os.path.join(os.path.dirname(self.file_path), new_name)
        os.rename(self.file_path, new_path)
        self.file_path = new_path

    def rename(self, old_sheet, new_sheet):
        if old_sheet in self.wb.sheetnames:
            self.wb[old_sheet].title = new_sheet

    def add_sheet(self, title):
        self.wb.create_sheet(title)

    def delete_sheet(self, name):
        if name in self.wb.sheetnames:
            sheet = self.wb[name]
            self.wb.remove(sheet)

    def sheet_exists(self, name):
        return name in self.wb.sheetnames

    def list_sheets(self):
        return self.wb.sheetnames

    def set_active(self, name):
        if name in self.wb.sheetnames:
            self.sheet = self.wb[name]

    def set_location(self, path):
        self.file_path = os.path.abspath(path)

    def get_location(self):
        return self.file_path

    def set_value(self, row, column, value):
        self.sheet.cell(row=row, column=column).value = value

    def get_value(self, row, column):
        return self.sheet.cell(row=row, column=column).value

    def replace(self, row, column, old, new):
        if self.sheet.cell(row=row, column=column).value == old:
            self.sheet.cell(row=row, column=column).value = new

    def replace_all(self, old, new):
        for row in self.sheet.iter_rows():
            for cell in row:
                if cell.value == old:
                    cell.value = new

    def clear_cell(self, row, column):
        self.sheet.cell(row=row, column=column).value = None

    def get_row(self, row_num):
        return [cell.value for cell in self.sheet[row_num]]

    def get_column(self, col_num):
        col_letter = get_column_letter(col_num)
        return [cell.value for cell in self.sheet[col_letter]]

    def mean(self, *, row: int = None, column: int = None) -> float:
        """
        Return the mean of a single row or single column.
        Usage: mean(row=2) or mean(column=3)
        """
        if row is not None:
            values = [v for v in self.get_row(row) if isinstance(v, (int, float))]
        elif column is not None:
            values = [v for v in self.get_column(column) if isinstance(v, (int, float))]
        else:
            raise ValueError("You must specify exactly one of row or column.")
        return statistics.mean(values) if values else 0.0

    def sum(self, *, row: int = None, column: int = None) -> float:
        """
        Return the sum of a single row or single column.
        Usage: sum(row=2) or sum(column=3)
        """
        if row is not None:
            values = [v for v in self.get_row(row) if isinstance(v, (int, float))]
        elif column is not None:
            values = [v for v in self.get_column(column) if isinstance(v, (int, float))]
        else:
            raise ValueError("You must specify exactly one of row or column.")
        return sum(values)

    def max(self, *, row: int = None, column: int = None) -> float:
        """
        Return the maximum value of a single row or single column.
        Usage: max(row=2) or max(column=3)
        """
        if row is not None:
            values = [v for v in self.get_row(row) if isinstance(v, (int, float))]
        elif column is not None:
            values = [v for v in self.get_column(column) if isinstance(v, (int, float))]
        else:
            raise ValueError("You must specify exactly one of row or column.")
        return max(values) if values else None

    def min(self, *, row: int = None, column: int = None) -> float:
        """
        Return the minimum value of a single row or single column.
        Usage: min(row=2) or min(column=3)
        """
        if row is not None:
            values = [v for v in self.get_row(row) if isinstance(v, (int, float))]
        elif column is not None:
            values = [v for v in self.get_column(column) if isinstance(v, (int, float))]
        else:
            raise ValueError("You must specify exactly one of row or column.")
        return min(values) if values else None

    def count(self, value):
        total = 0
        for row in self.sheet.iter_rows():
            for cell in row:
                if cell.value == value:
                    total += 1
        return total

    def get_dimensions(self):
        return self.sheet.max_row, self.sheet.max_column

    def display(self):
        data = []
        for row in self.sheet.iter_rows(values_only=True):
            data.append([cell if cell is not None else "" for cell in row])
        print(tabulate(data, tablefmt="grid"))

    def remove_column(self, column: int) -> None:
        """
        Remove the specified column from the active sheet.
        Args:
            column (int): 1-based column index to remove.
        Raises:
            ValueError: If column index is invalid or out of range.
        """
        max_col = self.sheet.max_column
        if not isinstance(column, int) or column < 1 or column > max_col:
            raise ValueError(f"Column index must be between 1 and {max_col}")
        self.sheet.delete_cols(column, 1)

    def remove_row(self, row: int) -> None:
        """
        Remove the specified row from the active sheet.
        Args:
            row (int): 1-based row index to remove.
        Raises:
            ValueError: If row index is invalid or out of Protocols:
            ValueError: If row index is invalid or out of range.
        """
        max_row = self.sheet.max_row
        if not isinstance(row, int) or row < 1 or row > max_row:
            raise ValueError(f"Row index must be between 1 and {max_row}")
        self.sheet.delete_rows(row, 1)

    def __repr__(self):
        return f"<ExcelFile at '{self.file_path}'>"